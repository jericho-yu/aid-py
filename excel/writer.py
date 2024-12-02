from typing import List, Any

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class Writer:
	def __init__(self, filename: str):
		if not filename:
			raise ValueError("Filename cannot be empty")
		self.filename = filename
		self.workbook = openpyxl.Workbook()
		self.sheet = self.workbook.active

	def get_filename(self) -> str:
		return self.filename

	def set_filename(self, filename: str) -> "Writer":
		self.filename = filename
		return self

	def create_sheet(self, sheet_name: str) -> "Writer":
		if not sheet_name:
			raise ValueError("Sheet name cannot be empty")
		self.sheet = self.workbook.create_sheet(title=sheet_name)
		return self

	def active_sheet_by_name(self, sheet_name: str) -> "Writer":
		if not sheet_name:
			raise ValueError("Sheet name cannot be empty")
		self.sheet = self.workbook[sheet_name]
		return self

	def active_sheet_by_index(self, sheet_index: int) -> "Writer":
		if sheet_index < 0:
			raise ValueError("Sheet index cannot be less than 0")
		self.sheet = self.workbook.worksheets[sheet_index]
		return self

	def set_sheet_name(self, sheet_name: str) -> "Writer":
		self.sheet.title = sheet_name
		return self

	def set_column_width_by_index(self, col: int, width: float) -> "Writer":
		self.set_columns_width_by_index(col, col, width)
		return self

	def set_column_width_by_text(self, col: str, width: float) -> "Writer":
		self.set_columns_width_by_text(col, col, width)
		return self

	def set_columns_width_by_index(
		self, start_col: int, end_col: int, width: float
	) -> "Writer":
		for col in range(start_col, end_col + 1):
			self.sheet.column_dimensions[get_column_letter(col)].width = width
		return self

	def set_columns_width_by_text(
		self, start_col: str, end_col: str, width: float
	) -> "Writer":
		for col in range(
			openpyxl.utils.column_index_from_string(start_col),
			openpyxl.utils.column_index_from_string(end_col) + 1,
		):
			self.sheet.column_dimensions[get_column_letter(col)].width = width
		return self

	def set_rows(self, rows: List[List[Any]]) -> "Writer":
		for row in rows:
			self.add_row(row)
		return self

	def add_row(self, row: List[Any]) -> "Writer":
		self.sheet.append(row)
		return self

	def set_title_row(self, titles: List[str], row_number: int) -> "Writer":
		for col_num, title in enumerate(titles, 1):
			cell = self.sheet.cell(row=row_number, column=col_num, value=title)
			cell.font = Font(bold=True)
		return self

	def save(self) -> None:
		if not self.filename:
			raise ValueError("Filename is not set")
		self.workbook.save(self.filename)

	def download(self, writer) -> None:
		writer.headers["Content-Type"] = "application/octet-stream"
		writer.headers["Content-Disposition"] = f"attachment; filename={self.filename}"
		self.workbook.save(writer)

	def get_workbook(self) -> openpyxl.Workbook:
		return self.workbook


if __name__ == "__main__":
	# Example usage
	writer = Writer("example.xlsx")
	writer.active_sheet_by_index(0).set_title_row(["Title1", "Title2"], 1).add_row(
		["Data1", "Data2"]
	).save()
